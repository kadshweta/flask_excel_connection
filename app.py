from flask import Flask ,request,jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI']='sqlite:///example.db'
db=SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Name=db.Column(db.String(200))
    Score=db.Column(db.Integer)

@app.route("/parse",methods=['POST'])
def show():

    data=request.files['file']
    workbook=load_workbook(data)
    sheet=workbook.active

    for row in sheet.iter_rows(values_only=True):
        
        data2=User(Name=row[0],Score=row[1])
        db.session.add(data2)
    
    db.session.commit()
    return jsonify({'msg':"done"})


if __name__=="__main__":
    with app.app_context():
        db.create_all()
        app.run(debug=True)






# import openpyxl as op
# file_path=op.load_workbook('C:\\excel\\Student_data.xlsx')
# print(file_path.sheetnames)
        



        # @app.route('/parse', methods=['POST'])
# def parse_excel():
#     try:
#         data = request.files['file']
#         workbook = load_workbook(data)
#         sheet = workbook.active

#         for row in sheet.iter_rows(min_row=2, values_only=True):            
#             try:
#                 data2=User(Name=row[0], Score=row[1])
#                 print(data2)
#                 db.session.add(data2)
#             except ValueError as e:
#                 print(f"Error unpacking row: {e}")
#         db.session.commit()

#         return jsonify({'message': 'ok'}), 200
#     except Exception as e:
#         return jsonify({'error': str(e)}), 500

